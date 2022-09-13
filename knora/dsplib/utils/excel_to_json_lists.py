"""This module handles all the operations which are used for the creation of JSON lists from Excel files."""
import glob
import json
import os
import re
import unicodedata
from typing import Any, Union, Optional, Tuple

import jsonschema
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
import regex

from knora.dsplib.models.helpers import BaseError

list_of_lists_of_previous_cell_values: list[list[str]] = []
"""Module level variable used to ensure that there are no duplicate node names"""

list_of_previous_node_names: list[str] = []
"""Module level variable used to ensure that there are no duplicate node names"""


def expand_lists_from_excel(
    lists_section: list[dict[str, Union[str, dict[str, Any]]]]
) -> Tuple[list[dict[str, Any]], bool]:
    """
    Checks if the "lists" section of a JSON project file contains references to Excel files. Expands all Excel files to
    JSON, and returns the expanded "lists" section. If there are no references to Excel files, the "lists" section is
    returned as is.
    Returns a tuple consisting of the expanded "lists" section and a boolean value: True if everything went smoothly,
    False if one of the lists couldn't be expanded correctly.

    Args:
        lists_section: the "lists" section of a parsed JSON project file. If this is an empty list, an empty list will be returned.

    Returns:
        the same "lists" section, but without references to Excel files
        True if all lists could be expanded correctly, False if a problem occurred
    """
    overall_success = True
    new_lists = []
    for _list in lists_section:
        if "folder" not in _list["nodes"]:
            # this list is a JSON list: return it as it is
            new_lists.append(_list)
        else:
            # this is a reference to a folder with Excel files
            excel_file_names = _extract_excel_file_names(_list["nodes"]["folder"])
            try:
                finished_list = _make_json_lists_from_excel(excel_file_names, verbose=False)
                new_lists.append(finished_list)
                print(f"\tThe list '{_list['name']}' contains a reference to the folder '{_list['nodes']['folder']}'. "
                      f"The Excel files therein will be temporarily expanded into the 'lists' section of your project.")
            except BaseError as err:
                print(f"\tWARNING: The list '{_list['name']}' contains a reference to the folder "
                      f"'{_list['nodes']['folder']}', but a problem occurred while trying to expand the Excel files "
                      f"therein into the 'lists' section of your project: {err.message}")
                overall_success = False

    return new_lists, overall_success


def _get_values_from_excel(
    excelfiles: dict[str, Worksheet],
    base_file: dict[str, Worksheet],
    parentnode: dict[str, Any],
    row: int,
    col: int,
    preval: list[str],
    verbose: bool = False
) -> tuple[int, dict[str, Any]]:
    """
    This function calls itself recursively to go through the Excel files. It extracts the cell values and composes
    the JSON list.

    Args:
        excelfiles: List of Excel files with the values in different languages
        base_file: File name of the base file
        parentnode: Name(s) of the parent node(s) of the current node
        row: The index of the current row of the Excel sheet
        col: The index of the current column of the Excel sheet
        preval: List of previous values, needed to check the consistency of the list hierarchy
        verbose: verbose switch

    Returns:
        int: Row index for the next loop (current row index minus 1)
        dict: The JSON list up to the current recursion. At the last recursion, this is the final JSON list.
    """
    nodes: list[dict[str, Any]] = []
    currentnode: dict[str, Any] = dict()
    base_file_ws: Worksheet = list(base_file.values())[0]
    cell: Cell = base_file_ws.cell(column=col, row=row)

    for excelfile in excelfiles.values():
        if any((not excelfile["A1"].value, excelfile["B1"].value)):
            raise BaseError(f"ERROR: Inconsistency in Excel list: The first row must consist of exactly one value, in "
                            f"cell A1. All other cells of row 1 must be empty.\nInstead, found the following:\n"
                            f"Cell A1: '{excelfile['A1'].value}'\n"
                            f"Cell B1: '{excelfile['B1'].value}'")

    if col > 1:
        # append the cell value of the parent node (which is one value to the left of the current cell) to the list of
        # previous values
        preval.append(base_file_ws.cell(column=col-1, row=row).value.strip())

    while cell.value and regex.search(r"\p{L}", cell.value, flags=re.UNICODE):
        # check if all predecessors in row (values to the left) are consistent with the values in preval list
        for idx, val in enumerate(preval[:-1]):
            if val != base_file_ws.cell(column=idx+1, row=row).value.strip():
                raise BaseError(f"ERROR: Inconsistency in Excel list: {val} not equal to "
                                f"{base_file_ws.cell(column=idx+1, row=row).value.strip()}")

        # loop through the row until the last (furthest right) value is found
        next_value = base_file_ws.cell(column=col+1, row=row).value
        if next_value and regex.search(r"\p{L}", next_value, flags=re.UNICODE):
            row, _ = _get_values_from_excel(
                excelfiles=excelfiles,
                base_file=base_file,
                parentnode=currentnode,
                col=col+1,
                row=row,
                preval=preval,
                verbose=verbose
            )

        # if value was last in row (no further values to the right), it's a node, continue here
        else:
            # check if there are duplicate nodes (i.e. identical rows), raise a BaseError if so
            new_check_list = preval.copy()
            new_check_list.append(cell.value.strip())
            list_of_lists_of_previous_cell_values.append(new_check_list)

            if _contains_duplicates(list_of_lists_of_previous_cell_values):
                raise BaseError(f"ERROR: There is at least one duplicate node in the list. Found duplicate in column "
                                f"{cell.column}, row {cell.row}:\n'{cell.value.strip()}'")

            # create a simplified version of the cell value and use it as name of the node
            nodename = simplify_name(cell.value.strip())
            list_of_previous_node_names.append(nodename)

            # append a number (p.ex. node-name-2) if there are list nodes with identical names
            n = list_of_previous_node_names.count(nodename)
            if n > 1:
                nodename = nodename + "-" + str(n)

            # read label values from the other Excel files (other languages)
            labels_dict: dict[str, str] = {}
            for other_lang, ws_other_lang in excelfiles.items():
                cell_value = ws_other_lang.cell(column=col, row=row).value
                if not(isinstance(cell_value, str) and len(cell_value) > 0):
                    raise BaseError(f"ERROR: Malformed Excel file: The Excel file with the language code "
                                    f"'{other_lang}' should have a value in row {row}, column {col}")
                else:
                    labels_dict[other_lang] = cell_value.strip()

            # create current node from extracted cell values and append it to the nodes list
            currentnode = {"name": nodename, "labels": labels_dict}
            nodes.append(currentnode)
            if verbose:
                print(f"Added list node: {cell.value.strip()} ({nodename})")

        # go one row down and repeat loop if there is a value
        row += 1
        cell = base_file_ws.cell(column=col, row=row)

    if col > 1:
        preval.pop()

    # add the new nodes to the parentnode
    parentnode["nodes"] = nodes

    return row - 1, parentnode


def _make_json_lists_from_excel(excel_file_names: list[str], verbose: bool = False) -> list[dict[str, Any]]:
    """
    Reads Excel files and transforms them into a dict structure which can later be inserted into the "lists" array
    of a JSON project file.

    Args:
        excel_file_names: Excel files to be processed
        verbose: verbose switch

    Returns:
        The finished "lists" section
    """
    # Define starting point in Excel file
    startrow = 1
    startcol = 1

    # Check if English file is available and take it as base file. Take last one from list of Excel files if English
    # is not available. The node names are later derived from the labels of the base file.
    base_file: dict[str, Worksheet] = dict()
    for filename in excel_file_names:
        if "_en.xlsx" in os.path.basename(filename):
            lang = "en"
            ws = load_workbook(filename, read_only=True).worksheets[0]
            base_file = {lang: ws}
    if len(base_file) == 0:
        file = excel_file_names[-1]
        lang = os.path.splitext(file)[0].split("_")[-1]
        ws = load_workbook(file, read_only=True).worksheets[0]
        base_file = {lang: ws}

    excelfiles: dict[str, Worksheet] = {}
    for f in excel_file_names:
        lang = os.path.splitext(f)[0].split("_")[-1]
        ws = load_workbook(f, read_only=True).worksheets[0]
        excelfiles[lang] = ws

    _, finished_lists = _get_values_from_excel(
        excelfiles=excelfiles,
        base_file=base_file,
        parentnode={},
        row=startrow,
        col=startcol,
        preval=[],
        verbose=verbose
    )

    # add the comments at the right place (comments are only mandatory for the root node of each list)
    for i, _list in enumerate(finished_lists["nodes"]):
        finished_lists["nodes"][i] = {
            "name": _list["name"],
            "labels": _list["labels"],
            "comments": _list["labels"],
            "nodes": _list["nodes"]
        }

    return finished_lists["nodes"]


def _contains_duplicates(list_to_check: list[Any]) -> bool:
    """
    Checks if the given list contains any duplicate items.

    Args:
        list_to_check: A list of items to be checked for duplicates

    Returns:
        True if there is a duplicate, false otherwise
    """
    has_duplicates = False

    for item in list_to_check:
        if list_to_check.count(item) > 1:
            has_duplicates = True

    return has_duplicates


def simplify_name(value: str) -> str:
    """
    Simplifies a given value in order to use it as node name

    Args:
        value: The value to be simplified

    Returns:
        str: The simplified value
    """
    simplified_value = str(value).lower()

    # normalize characters (p.ex. ä becomes a)
    simplified_value = unicodedata.normalize("NFKD", simplified_value)

    # replace forward slash and whitespace with a dash
    simplified_value = re.sub("[/\\s]+", "-", simplified_value)

    # delete all characters which are not letters, numbers or dashes
    simplified_value = re.sub("[^A-Za-z0-9\\-]+", "", simplified_value)

    return simplified_value


def validate_lists_section_with_schema(
    path_to_json_project_file: Optional[str] = None,
    lists_section: Optional[list[dict[str, Any]]] = None
) -> bool:
    """
    This function checks if a "lists" section of a JSON project is valid according to the schema. The "lists" section
    can be passed as path to the JSON project file, or as Python object. Only one of the two arguments should be passed.

    Args:
        path_to_json_project_file: path to the JSON project file that contains the "lists" section to validate
        lists_section: the "lists" section as Python object

    Returns:
        True if the list passed validation. Otherwise, a BaseError with a detailed error report is raised
    """
    assert bool(path_to_json_project_file) != bool(lists_section), "Exactly one of the two arguments must be given"
    with open("knora/dsplib/schemas/lists-only.json") as schema:
        lists_schema = json.load(schema)

    if path_to_json_project_file:
        with open(path_to_json_project_file) as f:
            project = json.load(f)
            lists_section = project["project"]["lists"]

    try:
        jsonschema.validate(instance={"lists": lists_section}, schema=lists_schema)
    except jsonschema.exceptions.ValidationError as err:
        raise BaseError(f'"Lists" section did not pass validation. The error message is: {err.message}\n'
                        f'The error occurred at {err.json_path}')
    return True


def _extract_excel_file_names(excelfolder: str) -> list[str]:
    """
    This method extracts the names of the Excel files that are in the folder, and asserts that they are named according
    to the requirements.

    Args:
        excelfolder: path to the folder containing the Excel file(s)

    Returns:
        list of the Excel file names to process
    """
    if not os.path.isdir(excelfolder):
        raise BaseError(f"ERROR: {excelfolder} is not a directory.")

    excel_file_names = [filename for filename in glob.iglob(f"{excelfolder}/*.xlsx")
                        if not os.path.basename(filename).startswith("~$")
                        and os.path.isfile(filename)]

    for filename in excel_file_names:
        if not re.search(r'.+_(de|en|fr|it|rm)\.xlsx$', filename):
            raise BaseError(f"Invalid filename '{excelfolder}/{filename}'. Expected format: ..._languagecode.xlsx")

    return excel_file_names


def list_excel2json(excelfolder: str, outfile: str) -> None:
    """
    This method writes a JSON file with a "lists" section that can later be inserted into a JSON project file.

    Args:
        excelfolder: path to the folder containing the Excel file(s)
        outfile: path to the JSON file the output is written into

    Returns:
        None
    """
    try:
        excel_file_names = _extract_excel_file_names(excelfolder)
        print("The following Excel files will be processed:")
        [print(f" - {filename}") for filename in excel_file_names]
        finished_lists = _make_json_lists_from_excel(excel_file_names, verbose=True)
        validate_lists_section_with_schema(lists_section=finished_lists)
    except BaseError as err:
        print(err.message)
        quit(1)

    with open(outfile, "w", encoding="utf-8") as fp:
        json.dump({"lists": finished_lists}, fp, indent=4, sort_keys=False, ensure_ascii=False)
        print("List was created successfully and written to file:", outfile)
