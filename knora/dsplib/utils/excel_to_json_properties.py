import json
import os
import re
from typing import Any

import jsonschema
from openpyxl import load_workbook


def validate_properties_with_schema(json_file: str) -> bool:
    """
    This function checks if the json properties are valid according to the schema.

    Args:
        json_file: the json with the properties to be validated

    Returns:
        True if the data passed validation, False otherwise

    """
    current_dir = os.path.dirname(os.path.realpath(__file__))
    with open(os.path.join(current_dir, '../schemas/properties-only.json')) as schema:
        properties_schema = json.load(schema)

    try:
        jsonschema.validate(instance=json_file, schema=properties_schema)
    except jsonschema.exceptions.ValidationError as err:
        print(err)
        return False
    print('Properties data passed schema validation.')
    return True


def properties_excel2json(excelfile: str, outfile: str) -> list[dict[str, Any]]:
    """
    Converts properties described in an Excel file into a properties section which can be integrated into a DSP ontology

    Args:
        excelfile: path to the Excel file containing the properties
        outfile: path to the output JSON file containing the properties section for the ontology

    Returns:
        List(JSON): a list with a dict (JSON) for each row in the Excel file
    """
    # load file
    wb = load_workbook(filename=excelfile, read_only=True)
    sheet = wb.worksheets[0]
    props = [row_to_prop(row) for row in sheet.iter_rows(min_row=2, values_only=True, max_col=13)
             if any(row) and any([re.search(r'[A-Za-z]+', elem) for elem in row if isinstance(elem, str)])]

    prefix = '"properties":'

    if validate_properties_with_schema(json.loads(json.dumps(props, indent=4))):
        # write final list to JSON file if list passed validation
        with open(file=outfile, mode='w+', encoding='utf-8') as file:
            file.write(prefix)
            json.dump(props, file, indent=4)
            print('Properties file was created successfully and written to file:', outfile)
    else:
        print('Properties data is not valid according to schema.')

    return props


def row_to_prop(row: tuple[str, str, str, str, str, str, str, str, str, str, str, str, str]) -> dict[str, Any]:
    """
    Parses the row of an Excel sheet and makes a property from it

    Args:
        row: the row of an Excel sheet

    Returns:
        prop (JSON): the property in JSON format
    """
    name, super_, object_, en, de, fr, it, comment_en, comment_de, comment_fr, comment_it, gui_element, gui_attributes = row
    labels = {}
    if en:
        labels['en'] = en.strip()
    if de:
        labels['de'] = de.strip()
    if fr:
        labels['fr'] = fr.strip()
    if it:
        labels['it'] = it.strip()
    if not labels:
        raise ValueError(f"No label given in any of the four languages: {name}")
    comments = {}
    if comment_en:
        comments['en'] = comment_en.strip()
    if comment_de:
        comments['de'] = comment_de.strip()
    if comment_fr:
        comments['fr'] = comment_fr.strip()
    if comment_it:
        comments['it'] = comment_it.strip()
    prop = {
        'name': name.strip(),
        'super': [elem.strip() for elem in super_.split(',')],
        'object': object_.strip(),
        'labels': labels,
        'comments': comments,
        'gui_element': gui_element.strip()
    }
    if gui_attributes:
        attr_list = [x.strip() for x in gui_attributes.split(',')]
        attr_dict = dict()
        for elem in attr_list:
            if ':' in elem:
                attr, val = [x.strip() for x in elem.split(':', maxsplit=1)]
                if re.search(r'\d+\.\d+', val):
                    val = float(val)
                elif re.search(r'\d+', val):
                    val = int(val)
                attr_dict.update({attr: val})
            elif object_.strip() == 'ListValue':
                attr_dict.update({'hlist': elem})
            else:
                raise ValueError(f'gui_attribute must be of the form "attr: value", except for ListValues, where the '
                                 f'simple name of the list is allowed. But the property "{name}", which is not a list, '
                                 f'has a gui_attribute that does not contain a colon.')
        prop['gui_attributes'] = attr_dict
    return prop
