import json
import os
from typing import Any

import jsonschema
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


def validate_resources_with_schema(json_file: str) -> bool:
    """
    This function checks if the json resources are valid according to the schema.

    Args:
        json_file: the json with the resources to be validated

    Returns:
        True if the data passed validation, False otherwise

    """
    current_dir = os.path.dirname(os.path.realpath(__file__))
    with open(os.path.join(current_dir, '../schemas/resources-only.json')) as schema:
        resources_schema = json.load(schema)

    try:
        jsonschema.validate(instance=json_file, schema=resources_schema)
    except jsonschema.exceptions.ValidationError as err:
        print(err)
        return False
    print('Resource data passed schema validation.')
    return True


def resources_excel2json(excelfile: str, outfile: str) -> None:
    """
    Converts properties described in an Excel file into a properties section which can be integrated into a DSP ontology

    Args:
        excelfile: path to the Excel file containing the properties
        outfile: path to the output JSON file containing the properties section for the ontology

    Returns:
        None
    """

    # load file
    wb = load_workbook(excelfile, read_only=True)

    # get overview
    sheet = wb['classes']
    resource_list = [c for c in sheet.iter_rows(min_row=2, values_only=True)]

    prefix = '"resources":'
    resources = [_extract_row(res, wb) for res in resource_list]

    if validate_resources_with_schema(json.loads(json.dumps(resources, indent=4))):
        # write final list to JSON file if list passed validation
        with open(file=outfile, mode='w+', encoding='utf-8') as file:
            file.write(prefix)
            json.dump(resources, file, indent=4)
            print('Resource file was created successfully and written to file:', outfile)
    else:
        print('Resource data is not valid according to schema.')


def _extract_row(row: tuple[str, str, str, str, str, str, str, str, str, str], wb: Workbook) -> dict[str, Any]:
    """build a property dict from a row of the excel file"""
    # get name
    name = row[0]
    # get labels
    labels = {}
    if row[1]:
        labels['en'] = row[1]
    if row[2]:
        labels['de'] = row[2]
    if row[3]:
        labels['fr'] = row[3]
    if row[4]:
        labels['it'] = row[4]
    # get comments
    comments = {}
    if row[5]:
        comments['en'] = row[5]
    if row[6]:
        comments['de'] = row[6]
    if row[7]:
        comments['fr'] = row[7]
    if row[8]:
        comments['it'] = row[8]
    # get super
    sup = row[9]

    # load details for this resource
    sh = wb[name]
    property_list = [c for c in sh.iter_rows(min_row=2, values_only=True)]

    cards = []
    # for each of the detail sheets
    for i, prop in enumerate(property_list):
        # get name and cardinality.
        # GUI-order is equal to order in the sheet.
        property_ = {
            "propname": ":" + prop[0],
            "cardinality": str(prop[1]),
            "gui_order": i + 1
        }
        cards.append(property_)

    # return resource dict
    return {
        "name": name,
        "labels": labels,
        "comments": comments,
        "super": sup,
        "cardinalities": cards
    }
