from typing import List, Set, Dict, Tuple, Optional
from openpyxl import load_workbook, worksheet

from ..models.connection import Connection
from ..models.project import Project
from ..models.listnode import ListNode
from ..models.langstring import Languages, LangString
from ..models.helpers import BaseError

from pprint import pprint

def list_creator(con: Connection, project: Project, parent_node: ListNode, nodes: List[dict]):
    nodelist = []
    for node in nodes:
        newnode = ListNode(
            con=con,
            project=project,
            label=node["labels"],
            comments=node.get("comments"),
            name=node["name"],
            parent=parent_node
        ).create()
        if node.get('nodes') is not None:
            subnodelist = list_creator(con, project, newnode, node['nodes'])
            nodelist.append({newnode.name: {"id": newnode.id, 'nodes': subnodelist}})
        else:
            nodelist.append({newnode.name: {"id": newnode.id}})
    return nodelist


def validate_list_from_excel(filepath: str,
                             sheetname: str,
                             startrow: int = 1,
                             startcol: int = 1,
                             verbose: bool = False):

    def analyse_level(ws: worksheet, level: int, row: int, col: int, preval: List[str]) ->int:
        cell = ws.cell(column=col, row=row)
        if col > 1:
            preval.append(ws.cell(column=col - 1, row=row).value)
        while cell.value:
            for idx, val in enumerate(preval[:-1]):
                if val != ws.cell(column=idx + 1, row=row).value:
                    raise BaseError(f"Inconsistency in Excel list! {val} not equal {ws.cell(column=idx + 1, row=row).value}")
            if ws.cell(column=col + 1, row=row).value:
                row = analyse_level(ws=ws, level=level + 1, col=col + 1, row=row, preval=preval)
                if not ws.cell(column=col, row=row).value:
                    if col > 1:
                        preval.pop()
                    return row
            else:
                if verbose:
                    print(f"Node on level={level}, value={cell.value} ok...")
            row += 1
            cell = ws.cell(column=col, row=row)
        if col > 1:
            preval.pop()
        return row - 1

    wb = load_workbook(filename=filepath, read_only=True)
    ws = wb[sheetname]
    tmp = []
    analyse_level(ws=ws, level=1, row=startrow, col=startcol, preval=tmp)


def json_list_from_excel(rootnode: {}, filepath: str, sheetname: str, startrow: int = 1, startcol: int = 1):

    names: Set[str] = set()

    def analyse_level(ws: worksheet, parentnode: {}, row: int, col: int, preval: List[str]) ->int:
        nodes: [] = []
        currentnode: {}

        cell = ws.cell(column=col, row=row)
        if col > 1:
            preval.append(ws.cell(column=col - 1, row=row).value)
        while cell.value:
            for idx, val in enumerate(preval[:-1]):
                if val != ws.cell(column=idx + 1, row=row).value:
                    raise BaseError(f"Inconsistency in Excel list! {val} not equal {ws.cell(column=idx + 1, row=row).value}")
            if ws.cell(column=col + 1, row=row).value:
                row = analyse_level(ws=ws, parentnode=currentnode, col=col + 1, row=row, preval=preval)
                if not ws.cell(column=col, row=row).value:
                    if col > 1:
                        preval.pop()
                    parentnode["nodes"] = nodes
                    return row
            else:
                tmpstr = cell.value
                tmpstr = tmpstr.split(" ")
                tmpstr = [w.title() for w in tmpstr]
                tmpstr = "".join(tmpstr)
                tmpstr = tmpstr[0].lower() + tmpstr[1:]
                while tmpstr in names:
                    tmpstr = tmpstr + "_"
                names.add(tmpstr)
                currentnode = {
                    "name": tmpstr,
                    "labels": {"en": cell.value}
                }
                nodes.append(currentnode)
                print(f"Adding list node: value={cell.value}")
            row += 1
            cell = ws.cell(column=col, row=row)
        if col > 1:
            preval.pop()
        parentnode["nodes"] = nodes
        return row - 1

    wb = load_workbook(filename=filepath, read_only=True)
    ws = wb[sheetname]
    tmp = []
    analyse_level(ws=ws, parentnode=rootnode, row=startrow, col=startcol, preval=tmp)


