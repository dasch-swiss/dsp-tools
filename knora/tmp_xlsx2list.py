import argparse
import sys
import os
import json

from typing import List, Set, Dict, Tuple, Optional
from openpyxl import load_workbook, worksheet


def json_list_from_excel(rootnode: {}, filepath: str, sheetname: str, startrow: int = 1, startcol: int = 1):

    def analyse_level(ws: worksheet, parentnode: {}, row: int, col: int, preval: List[str]) ->int:
        nodes: [] = []
        currentnode: {}

        cell = ws.cell(column=col, row=row)
        if col > 1:
            preval.append(ws.cell(column=col - 1, row=row).value)
        while cell.value:
            for idx, val in enumerate(preval[:-1]):
                if val != ws.cell(column=idx + 1, row=row).value:
                    raise BaseError(f"Inconsistency in Excel list! {val} not equal {ws.cell(column=idx + 1, row=row).value}")
            if ws.cell(column=col + 1, row=row).value:
                row = analyse_level(ws=ws, parentnode=currentnode, col=col + 1, row=row, preval=preval)
                if not ws.cell(column=col, row=row).value:
                    if col > 1:
                        preval.pop()
                    parentnode["nodes"] = nodes
                    return row
            else:
                currentnode = {
                    "name": f"node_{row}_{col}",
                    "labels": {"en": cell.value}
                }
                nodes.append(currentnode)
                print(f"Adding list node: value={cell.value}")
            row += 1
            cell = ws.cell(column=col, row=row)
        if col > 1:
            preval.pop()
        parentnode["nodes"] = nodes
        return row - 1

    wb = load_workbook(filename=filepath, read_only=True)
    ws = wb[sheetname]
    tmp = []
    analyse_level(ws=ws, parentnode=rootnode, row=startrow, col=startcol, preval=tmp)


def program(args):
    #
    # parse the arguments of the command line
    #
    parser = argparse.ArgumentParser(
        description="A program to create and manipulate ontologies based on the DaSCH Service Platform"
    )
    parser.add_argument("-j", "--jsonfile", type=str, help="Filename of JSON file produced", default="list.json")
    parser.add_argument("-S", "--sheet", type=str, help="Name of excel sheet to be used", default="Tabelle1")
    parser.add_argument("-l", "--listname", type=str, help="Name of list to be created", default="my_list")
    parser.add_argument("-L", "--label", type=str, help="Label of list to be created", default="MyList")
    parser.add_argument("-x", "--lang", type=str, help="Language for label", default="en")
    parser.add_argument("-v", "--verbose", action="store_true", help="Verbose feedback")
    parser.add_argument("excelfile", help="Path to the excel file containing the list data", default="list.xlsx")
    parser.add_argument("outfile", help="Path to the output JSON file containing the list data", default="list.json")

    args = parser.parse_args()

    rootnode = {
        "name": args.listname,
        "labels": {args.lang: args.label}
    }

    json_list_from_excel(rootnode, args.excelfile, args.sheet, 1, 1)

    with open(args.outfile, 'w', encoding="utf-8") as fp:
        json.dump(rootnode, fp, indent=3, sort_keys=True)


def main():
    program(sys.argv[1:])


if __name__ == '__main__':
    program(sys.argv[1:])
