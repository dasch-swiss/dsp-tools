"""This module handles all the operations which are used for the creation of JSON lists from Excel files."""

import importlib.resources
import json
from pathlib import Path
from typing import Any
from typing import Optional
from typing import cast

import jsonschema
import regex
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from dsp_tools.commands.excel2json.models.input_error import MoreThanOneSheetProblem
from dsp_tools.commands.excel2json.models.list_node_name import ListNodeNames
from dsp_tools.error.exceptions import BaseError
from dsp_tools.error.exceptions import InputError
from dsp_tools.utils.data_formats.shared import simplify_name


def old_excel2lists(
    excelfolder: str,
    path_to_output_file: Optional[str] = None,
    verbose: bool = False,
) -> tuple[list[dict[str, Any]], bool]:
    """
    Converts lists described in Excel files into a "lists" section that can be inserted into a JSON project file.

    Args:
        excelfolder: path to the folder containing the Excel file(s)
        path_to_output_file: if provided, the output is written into this JSON file
        verbose: verbose switch

    Raises:
        InputError: if something went wrong
        BaseError: if something went wrong

    Returns:
        a tuple consisting of the "lists" section as Python list, and the success status (True if everything went well)
    """
    excel_file_paths = _extract_excel_file_paths(excelfolder)
    if verbose:
        print("The following Excel files will be processed:")
        print(*(f" - {filename}" for filename in excel_file_paths), sep="\n")

    finished_lists = _make_json_lists_from_excel(excel_file_paths, verbose=verbose)
    validate_lists_section_with_schema(lists_section=finished_lists)

    if path_to_output_file:
        with open(path_to_output_file, "w", encoding="utf-8") as fp:
            json.dump(finished_lists, fp, indent=4, ensure_ascii=False)
            print(f"lists section was created successfully and written to file '{path_to_output_file}'")

    return finished_lists, True


def _get_values_from_excel(
    excelfiles: dict[str, Worksheet],
    base_file: dict[str, Worksheet],
    parentnode: dict[str, Any],
    row: int,
    col: int,
    preval: list[str],
    list_node_names: ListNodeNames,
    verbose: bool = False,
) -> tuple[int, dict[str, Any]]:
    """
    This function calls itself recursively to go through the Excel files. It extracts the cell values and composes
    the JSON list.

    Args:
        excelfiles: List of Excel files with the values in different languages
        base_file: File name of the base file
        parentnode: Name(s) of the parent node(s) of the current node
        row: The index of the current row of the Excel sheet
        col: The index of the current column of the Excel sheet
        preval: List of previous values, needed to check the consistency of the list hierarchy
        list_node_names: object containing the information which node names were used by previous lists
        verbose: verbose switch

    Raises:
        InputError: if one of the Excel files contains invalid data

    Returns:
        int: Row index for the next loop (current row index minus 1)
        dict: The JSON list up to the current recursion. At the last recursion, this is the final JSON list.
    """
    nodes: list[dict[str, Any]] = []
    currentnode: dict[str, Any] = {}
    base_file_ws: Worksheet = next(iter(base_file.values()))
    cell: Cell = cast(Cell, base_file_ws.cell(column=col, row=row))

    for excelfile in excelfiles.values():
        if any((not excelfile["A1"].value, excelfile["B1"].value)):
            raise InputError(
                f"ERROR: Inconsistency in Excel list: The first row must consist of exactly one value, in cell A1. "
                f"All other cells of row 1 must be empty.\nInstead, found the following:\n"
                f" - Cell A1: '{excelfile['A1'].value}'\n"
                f" - Cell B1: '{excelfile['B1'].value}'"
            )

    if col > 1:
        # append the cell value of the parent node (which is one value to the left of the current cell) to the list of
        # previous values
        preval.append(str(base_file_ws.cell(column=col - 1, row=row).value).strip())

    while cell.value and regex.search(r"\p{L}", str(cell.value), flags=regex.UNICODE):
        _check_if_predecessors_in_row_are_consistent_with_preval(base_file_ws, preval, row)

        # loop through the row until the last (furthest right) value is found
        next_value = base_file_ws.cell(column=col + 1, row=row).value
        if next_value and regex.search(r"\p{L}", str(next_value), flags=regex.UNICODE):
            row, _ = _get_values_from_excel(
                excelfiles=excelfiles,
                base_file=base_file,
                parentnode=currentnode,
                col=col + 1,
                row=row,
                preval=preval,
                list_node_names=list_node_names,
                verbose=verbose,
            )

        # if value was last in row (no further values to the right), it's a node, continue here
        else:
            currentnode = _make_new_node(cell, col, excelfiles, preval, row, list_node_names, verbose)
            nodes.append(currentnode)

        # go one row down and repeat loop if there is a value
        row += 1
        cell = cast(Cell, base_file_ws.cell(column=col, row=row))

    if col > 1:
        preval.pop()

    # add the new nodes to the parentnode
    parentnode["nodes"] = nodes

    return row - 1, parentnode


def _check_if_predecessors_in_row_are_consistent_with_preval(
    base_file_ws: Worksheet, preval: list[str], row: int
) -> None:
    for idx, val in enumerate(preval[:-1]):
        if val != str(base_file_ws.cell(column=idx + 1, row=row).value).strip():
            raise InputError(
                "ERROR: Inconsistency in Excel list: "
                f"{val} not equal to {str(base_file_ws.cell(column=idx + 1, row=row).value).strip()}"
            )


def _make_new_node(
    cell: Cell,
    col: int,
    excelfiles: dict[str, Worksheet],
    preval: list[str],
    row: int,
    list_node_names: ListNodeNames,
    verbose: bool = False,
) -> dict[str, Any]:
    _check_if_duplicate_nodes_exist(cell, list_node_names, preval)
    # create a simplified version of the cell value and use it as name of the node
    nodename = simplify_name(str(cell.value).strip())
    list_node_names.previous_node_names.append(nodename)
    # append a number (p.ex. node-name-2) if there are list nodes with identical names
    n = list_node_names.previous_node_names.count(nodename)

    if n > 1:
        nodename = f"{nodename}-{n}"

    labels_dict = _get_all_languages_of_node(excelfiles, col, row)

    # create the current node from extracted cell values and append it to the nodes list
    currentnode = {"name": nodename, "labels": labels_dict}
    if verbose:
        print(f"Added list node: {str(cell.value).strip()} ({nodename})")
    return currentnode


def _check_if_duplicate_nodes_exist(cell: Cell, list_node_names: ListNodeNames, preval: list[str]) -> None:
    new_check_list = preval.copy()
    new_check_list.append(str(cell.value).strip())
    list_node_names.lists_with_previous_cell_values.append(new_check_list)
    if any(
        list_node_names.lists_with_previous_cell_values.count(x) > 1
        for x in list_node_names.lists_with_previous_cell_values
    ):
        raise InputError(
            f"ERROR: There is at least one duplicate node in the list. "
            f"Found duplicate in column {cell.column}, row {cell.row}:\n'{str(cell.value).strip()}'"
        )


def _get_all_languages_of_node(excelfiles: dict[str, Worksheet], col: int, row: int) -> dict[str, str]:
    labels_dict: dict[str, str] = {}
    for other_lang, ws_other_lang in excelfiles.items():
        cell_value = ws_other_lang.cell(column=col, row=row).value
        if not (isinstance(cell_value, str) and len(cell_value) > 0):
            raise InputError(
                "ERROR: Malformed Excel file: The Excel file with the language code "
                f"'{other_lang}' should have a value in row {row}, column {col}"
            )
        else:
            labels_dict[other_lang] = cell_value.strip()
    return labels_dict


def _make_json_lists_from_excel(
    excel_file_paths: list[Path],
    verbose: bool = False,
) -> list[dict[str, Any]]:
    """
    Reads Excel files and transforms them into a list of dictionaries that can be used as "lists" array of a JSON
    project file.

    Args:
        excel_file_paths: Excel files to be processed
        verbose: verbose switch

    Raises:
        InputError: if one of the Excel files contains invalid data

    Returns:
        The finished "lists" section
    """

    # Define starting point in Excel file
    startrow = 1
    startcol = 1

    lang_to_worksheet: dict[str, Worksheet] = {x.stem: _read_and_check_workbook(x) for x in excel_file_paths}

    base_lang = "en" if "en" in lang_to_worksheet else next(iter(lang_to_worksheet.keys()))
    base_file = {base_lang: lang_to_worksheet[base_lang]}

    list_node_names = ListNodeNames()

    # construct the entire "lists" section as children of a fictive dummy parent node
    _, _list = _get_values_from_excel(
        excelfiles=lang_to_worksheet,
        base_file=base_file,
        parentnode={},
        row=startrow,
        col=startcol,
        preval=[],
        list_node_names=list_node_names,
        verbose=verbose,
    )

    # extract the children of the fictive dummy parent node
    finished_lists: list[dict[str, Any]] = _list["nodes"]

    # the "comments" section is mandatory for the root node of each list, so make a copy of "labels" and insert it at
    # the right place
    for i, _list in enumerate(finished_lists):
        finished_lists[i] = {
            "name": _list["name"],
            "labels": _list["labels"],
            "comments": _list["labels"],
            "nodes": _list["nodes"],
        }

    return finished_lists


def _read_and_check_workbook(excelpath: Path) -> Worksheet:
    all_worksheets = cast(list[Worksheet], load_workbook(excelpath, read_only=True).worksheets)
    if len(all_worksheets) != 1:
        msg = MoreThanOneSheetProblem(excelpath.name, [x.title for x in all_worksheets]).execute_error_protocol()
        raise InputError(msg)
    return all_worksheets[0]


def validate_lists_section_with_schema(
    path_to_json_project_file: Optional[str] = None,
    lists_section: Optional[list[dict[str, Any]]] = None,
) -> bool:
    """
    This function checks if a "lists" section of a JSON project is valid according to the schema. The "lists" section
    can be passed as path to the JSON project file, or as Python object. Only one of the two arguments should be passed.

    Args:
        path_to_json_project_file: path to the JSON project file that contains the "lists" section to validate
        lists_section: the "lists" section as Python object

    Raises:
        InputError: if the validation fails
        BaseError: if this function is called with invalid parameters

    Returns:
        True if the "lists" section passed validation
    """
    err_msg = "Validation of the 'lists' section works only if exactly one of the two arguments is given."
    match path_to_json_project_file, lists_section:
        case None, None:
            raise BaseError(err_msg)
        case str(), list():
            raise BaseError(err_msg)

    with (
        importlib.resources.files("dsp_tools")
        .joinpath("resources/schema/lists-only.json")
        .open(encoding="utf-8") as schema_file
    ):
        lists_schema = json.load(schema_file)

    if path_to_json_project_file:
        with open(path_to_json_project_file, encoding="utf-8") as f:
            project = json.load(f)
            lists_section = project["project"].get("lists")
            if not lists_section:
                raise InputError(
                    f"Cannot validate 'lists' section of {path_to_json_project_file}, "
                    "because there is no 'lists' section in this file."
                )

    try:
        jsonschema.validate(instance={"lists": lists_section}, schema=lists_schema)
    except jsonschema.ValidationError as err:
        raise InputError(
            f"'lists' section did not pass validation. The error message is: {err.message}\n"
            f"The error occurred at {err.json_path}"
        ) from None

    return True


def _extract_excel_file_paths(excelfolder: str) -> list[Path]:
    """
    This method extracts the names of the Excel files that are in the folder, and asserts that they are named according
    to the requirements.

    Args:
        excelfolder: path to the folder containing the Excel file(s)

    Raises:
        InputError: if excelfolder is not a directory, or if one of the files in it has an invalid name

    Returns:
        list of the Excel file paths to process
    """
    if not Path(excelfolder).is_dir():
        raise InputError(f"ERROR: {excelfolder} is not a directory.")

    supported_files = ["en.xlsx", "de.xlsx", "fr.xlsx", "it.xlsx", "rm.xlsx"]
    excel_file_paths = [x for x in Path(excelfolder).glob("*.xlsx") if x.is_file() and not x.name.startswith("~$")]

    for filepath in excel_file_paths:
        if filepath.name not in supported_files:
            raise InputError(f"Invalid file name '{filepath}'. Expected format: 'languagecode.xlsx'")

    return excel_file_paths
