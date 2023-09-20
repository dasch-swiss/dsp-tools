"""This module handles all the operations which are used for the creation of JSON lists from Excel files."""
import glob
import importlib.resources
import json
import os
from typing import Any, Optional, Union

import jsonschema
import regex
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from dsp_tools.models.exceptions import BaseError, UserError
from dsp_tools.utils.shared import simplify_name

list_of_lists_of_previous_cell_values: list[list[str]] = []
"""Module level variable used to ensure that there are no duplicate node names"""

list_of_previous_node_names: list[str] = []
"""Module level variable used to ensure that there are no duplicate node names"""


def expand_lists_from_excel(
    lists_section: list[dict[str, Union[str, dict[str, Any]]]],
) -> list[dict[str, Any]]:
    """
    Checks if the "lists" section of a JSON project file contains references to Excel files.
    Expands all Excel files to JSON,
    and returns the expanded "lists" section.
    If there are no references to Excel files,
    the "lists" section is returned as is.

    Args:
        lists_section: the "lists" section of a parsed JSON project file.
            If this is an empty list, an empty list will be returned.

    Raises:
        UserError: if a problem occurred while trying to expand the Excel files

    Returns:
        the same "lists" section, but without references to Excel files
    """
    new_lists = []
    for _list in lists_section:
        # case 1: this list is a JSON list: return it as it is
        if "folder" not in _list["nodes"]:
            new_lists.append(_list)
            continue

        # case 2: this is a reference to a folder with Excel files
        foldername = _list["nodes"]["folder"]  # type: ignore[index]  # types are too complex to annotate them correctly
        excel_file_paths = _extract_excel_file_paths(foldername)
        try:
            returned_lists_section = _make_json_lists_from_excel(excel_file_paths, verbose=False)
            # we only need the "nodes" section of the first element of the returned "lists" section. This "nodes"
            # section needs to replace the Excel folder reference. But the rest of the user-defined list element
            # needs to stay intact, e.g. the labels and comments.
            _list["nodes"] = returned_lists_section[0]["nodes"]
            new_lists.append(_list)
            print(
                f"\tThe list '{_list['name']}' contains a reference to the folder '{foldername}'. The Excel "
                f"files therein have been temporarily expanded into the 'lists' section of your project."
            )
        except BaseError as err:
            raise UserError(
                f"\tWARNING: The list '{_list['name']}' contains a reference to the folder '{foldername}', but a "
                f"problem occurred while trying to expand the Excel files therein into the 'lists' section of "
                f"your project: {err.message}"
            ) from None

    return new_lists


def _get_values_from_excel(
    excelfiles: dict[str, Worksheet],
    base_file: dict[str, Worksheet],
    parentnode: dict[str, Any],
    row: int,
    col: int,
    preval: list[str],
    verbose: bool = False,
) -> tuple[int, dict[str, Any]]:
    """
    This function calls itself recursively to go through the Excel files. It extracts the cell values and composes
    the JSON list.

    Args:
        excelfiles: List of Excel files with the values in different languages
        base_file: File name of the base file
        parentnode: Name(s) of the parent node(s) of the current node
        row: The index of the current row of the Excel sheet
        col: The index of the current column of the Excel sheet
        preval: List of previous values, needed to check the consistency of the list hierarchy
        verbose: verbose switch

    Raises:
        UserError: if one of the Excel files contains invalid data

    Returns:
        int: Row index for the next loop (current row index minus 1)
        dict: The JSON list up to the current recursion. At the last recursion, this is the final JSON list.
    """
    nodes: list[dict[str, Any]] = []
    currentnode: dict[str, Any] = dict()
    base_file_ws: Worksheet = list(base_file.values())[0]
    cell: Cell = base_file_ws.cell(column=col, row=row)

    for excelfile in excelfiles.values():
        if any((not excelfile["A1"].value, excelfile["B1"].value)):
            raise UserError(
                f"ERROR: Inconsistency in Excel list: The first row must consist of exactly one value, in cell A1. "
                f"All other cells of row 1 must be empty.\nInstead, found the following:\n"
                f" - Cell A1: '{excelfile['A1'].value}'\n"
                f" - Cell B1: '{excelfile['B1'].value}'"
            )

    if col > 1:
        # append the cell value of the parent node (which is one value to the left of the current cell) to the list of
        # previous values
        preval.append(str(base_file_ws.cell(column=col - 1, row=row).value).strip())

    while cell.value and regex.search(r"\p{L}", str(cell.value), flags=regex.UNICODE):
        # check if all predecessors in row (values to the left) are consistent with the values in preval list
        for idx, val in enumerate(preval[:-1]):
            if val != str(base_file_ws.cell(column=idx + 1, row=row).value).strip():
                raise UserError(
                    "ERROR: Inconsistency in Excel list: "
                    f"{val} not equal to {str(base_file_ws.cell(column=idx+1, row=row).value).strip()}"
                )

        # loop through the row until the last (furthest right) value is found
        next_value = base_file_ws.cell(column=col + 1, row=row).value
        if next_value and regex.search(r"\p{L}", str(next_value), flags=regex.UNICODE):
            row, _ = _get_values_from_excel(
                excelfiles=excelfiles,
                base_file=base_file,
                parentnode=currentnode,
                col=col + 1,
                row=row,
                preval=preval,
                verbose=verbose,
            )

        # if value was last in row (no further values to the right), it's a node, continue here
        else:
            # check if there are duplicate nodes (i.e. identical rows), raise a UserError if so
            new_check_list = preval.copy()
            new_check_list.append(str(cell.value).strip())
            list_of_lists_of_previous_cell_values.append(new_check_list)

            if any(list_of_lists_of_previous_cell_values.count(x) > 1 for x in list_of_lists_of_previous_cell_values):
                raise UserError(
                    f"ERROR: There is at least one duplicate node in the list. "
                    f"Found duplicate in column {cell.column}, row {cell.row}:\n'{str(cell.value).strip()}'"
                )

            # create a simplified version of the cell value and use it as name of the node
            nodename = simplify_name(str(cell.value).strip())
            list_of_previous_node_names.append(nodename)

            # append a number (p.ex. node-name-2) if there are list nodes with identical names
            n = list_of_previous_node_names.count(nodename)
            if n > 1:
                nodename = nodename + "-" + str(n)

            # read label values from the other Excel files (other languages)
            labels_dict: dict[str, str] = {}
            for other_lang, ws_other_lang in excelfiles.items():
                cell_value = ws_other_lang.cell(column=col, row=row).value
                if not (isinstance(cell_value, str) and len(cell_value) > 0):
                    raise UserError(
                        "ERROR: Malformed Excel file: The Excel file with the language code "
                        f"'{other_lang}' should have a value in row {row}, column {col}"
                    )
                else:
                    labels_dict[other_lang] = cell_value.strip()

            # create current node from extracted cell values and append it to the nodes list
            currentnode = {"name": nodename, "labels": labels_dict}
            nodes.append(currentnode)
            if verbose:
                print(f"Added list node: {str(cell.value).strip()} ({nodename})")

        # go one row down and repeat loop if there is a value
        row += 1
        cell = base_file_ws.cell(column=col, row=row)

    if col > 1:
        preval.pop()

    # add the new nodes to the parentnode
    parentnode["nodes"] = nodes

    return row - 1, parentnode


def _make_json_lists_from_excel(
    excel_file_paths: list[str],
    verbose: bool = False,
) -> list[dict[str, Any]]:
    """
    Reads Excel files and transforms them into a list of dictionaries that can be used as "lists" array of a JSON
    project file.

    Args:
        excel_file_paths: Excel files to be processed
        verbose: verbose switch

    Raises:
        UserError: if one of the Excel files contains invalid data

    Returns:
        The finished "lists" section
    """
    # reset the global variables
    global list_of_previous_node_names
    global list_of_lists_of_previous_cell_values
    list_of_previous_node_names = []
    list_of_lists_of_previous_cell_values = []

    # Define starting point in Excel file
    startrow = 1
    startcol = 1

    # make a dict with the language labels and the worksheets
    lang_to_worksheet: dict[str, Worksheet] = {}
    for filepath in excel_file_paths:
        lang_to_worksheet[os.path.basename(filepath)[0:2]] = load_workbook(filepath, read_only=True).worksheets[0]

    # take English as base file. If English is not available, take a random one.
    base_lang = "en" if "en" in lang_to_worksheet else list(lang_to_worksheet.keys())[0]
    base_file = {base_lang: lang_to_worksheet[base_lang]}

    # construct the entire "lists" section as children of a fictive dummy parent node
    _, _list = _get_values_from_excel(
        excelfiles=lang_to_worksheet,
        base_file=base_file,
        parentnode={},
        row=startrow,
        col=startcol,
        preval=[],
        verbose=verbose,
    )

    # extract the children of the fictive dummy parent node
    finished_lists: list[dict[str, Any]] = _list["nodes"]

    # the "comments" section is mandatory for the root node of each list, so make a copy of "labels" and insert it at
    # the right place
    for i, _list in enumerate(finished_lists):
        finished_lists[i] = {
            "name": _list["name"],
            "labels": _list["labels"],
            "comments": _list["labels"],
            "nodes": _list["nodes"],
        }

    return finished_lists


def validate_lists_section_with_schema(
    path_to_json_project_file: Optional[str] = None,
    lists_section: Optional[list[dict[str, Any]]] = None,
) -> bool:
    """
    This function checks if a "lists" section of a JSON project is valid according to the schema. The "lists" section
    can be passed as path to the JSON project file, or as Python object. Only one of the two arguments should be passed.

    Args:
        path_to_json_project_file: path to the JSON project file that contains the "lists" section to validate
        lists_section: the "lists" section as Python object

    Raises:
        UserError: if the validation fails
        BaseError: if this function is called with invalid parameters

    Returns:
        True if the "lists" section passed validation
    """
    if bool(path_to_json_project_file) == bool(lists_section):
        raise BaseError("Validation of the 'lists' section works only if exactly one of the two arguments is given.")

    with importlib.resources.files("dsp_tools").joinpath("resources/schema/lists-only.json").open(
        encoding="utf-8"
    ) as schema_file:
        lists_schema = json.load(schema_file)

    if path_to_json_project_file:
        with open(path_to_json_project_file, encoding="utf-8") as f:
            project = json.load(f)
            lists_section = project["project"].get("lists")
            if not lists_section:
                raise UserError(
                    f"Cannot validate 'lists' section of {path_to_json_project_file}, "
                    "because there is no 'lists' section in this file."
                )

    try:
        jsonschema.validate(instance={"lists": lists_section}, schema=lists_schema)
    except jsonschema.ValidationError as err:
        raise UserError(
            f"'lists' section did not pass validation. The error message is: {err.message}\n"
            f"The error occurred at {err.json_path}"
        ) from None

    return True


def _extract_excel_file_paths(excelfolder: str) -> list[str]:
    """
    This method extracts the names of the Excel files that are in the folder, and asserts that they are named according
    to the requirements.

    Args:
        excelfolder: path to the folder containing the Excel file(s)

    Raises:
        UserError: if excelfolder is not a directory, or if one of the files in it has an invalid name

    Returns:
        list of the Excel file paths to process
    """
    if not os.path.isdir(excelfolder):
        raise UserError(f"ERROR: {excelfolder} is not a directory.")

    excel_file_paths = [
        filename
        for filename in glob.iglob(f"{excelfolder}/*.xlsx")
        if not os.path.basename(filename).startswith("~$") and os.path.isfile(filename)
    ]

    for filepath in excel_file_paths:
        if not regex.search(r"^(de|en|fr|it|rm)\.xlsx$", os.path.basename(filepath)):
            raise UserError(f"Invalid file name '{filepath}'. Expected format: 'languagecode.xlsx'")

    return excel_file_paths


def excel2lists(
    excelfolder: str,
    path_to_output_file: Optional[str] = None,
    verbose: bool = False,
) -> tuple[list[dict[str, Any]], bool]:
    """
    Converts lists described in Excel files into a "lists" section that can be inserted into a JSON project file.

    Args:
        excelfolder: path to the folder containing the Excel file(s)
        path_to_output_file: if provided, the output is written into this JSON file
        verbose: verbose switch

    Raises:
        UserError: if something went wrong
        BaseError: if something went wrong

    Returns:
        a tuple consisting of the "lists" section as Python list, and the success status (True if everything went well)
    """
    # read the data
    excel_file_paths = _extract_excel_file_paths(excelfolder)
    if verbose:
        print("The following Excel files will be processed:")
        print(*(f" - {filename}" for filename in excel_file_paths), sep="\n")

    # construct the "lists" section
    finished_lists = _make_json_lists_from_excel(excel_file_paths, verbose=verbose)
    validate_lists_section_with_schema(lists_section=finished_lists)

    # write final "lists" section
    if path_to_output_file:
        with open(path_to_output_file, "w", encoding="utf-8") as fp:
            json.dump(finished_lists, fp, indent=4, ensure_ascii=False)
            print('"lists" section was created successfully and written to file:', path_to_output_file)

    return finished_lists, True
