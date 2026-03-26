import urllib.parse
from pathlib import Path

import openpyxl
import pytest
import requests

from dsp_tools.cli.args import ServerCredentials
from dsp_tools.commands.mapping.config_file import parse_mapping_config
from dsp_tools.commands.mapping.mapping_add import mapping_add
from dsp_tools.utils.data_formats.iri_util import make_dsp_ontology_prefix

SHORTCODE = "4125"
ONTO_NAME = "e2e-testonto"
CLASS_LOCAL_NAME = "ImageResource"
PROP_LOCAL_NAME = "hasText"


@pytest.fixture(scope="module")
def ontology_iri(creds: ServerCredentials) -> str:
    return make_dsp_ontology_prefix(creds.server, SHORTCODE, ONTO_NAME).rstrip("#")


@pytest.fixture(scope="module")
def excel_path(tmp_path_factory: pytest.TempPathFactory) -> Path:
    path = tmp_path_factory.mktemp("mapping") / "mappings.xlsx"
    wb = openpyxl.Workbook()

    prefix_ws = wb.active
    assert prefix_ws is not None
    prefix_ws.title = "prefix"
    prefix_ws.append(["prefix", "link"])
    prefix_ws.append(["schema", "http://schema.org/"])

    classes_ws = wb.create_sheet("classes")
    classes_ws.append(["class", "mapping"])
    classes_ws.append([CLASS_LOCAL_NAME, "schema:Thing"])

    props_ws = wb.create_sheet("properties")
    props_ws.append(["property", "mapping"])
    props_ws.append([PROP_LOCAL_NAME, "schema:text"])

    wb.save(path)
    return path


@pytest.fixture(scope="module")
def config_path(tmp_path_factory: pytest.TempPathFactory, creds: ServerCredentials, excel_path: Path) -> Path:
    path = tmp_path_factory.mktemp("mapping_cfg") / f"mapping-{SHORTCODE}-{ONTO_NAME}.yaml"
    path.write_text(
        f"shortcode: '{SHORTCODE}'\n"
        f"ontology: {ONTO_NAME}\n"
        f"server: {creds.server}\n"
        f"user: {creds.user}\n"
        f"password: {creds.password}\n"
        f"excel-file: {excel_path}\n",
        encoding="utf-8",
    )
    return path


@pytest.mark.xfail(reason="v3 mapping endpoints may not yet be in the testcontainer image", strict=False)
def test_mapping_add_success(config_path: Path, ontology_iri: str, creds: ServerCredentials) -> None:
    info = parse_mapping_config(config_path)
    result = mapping_add(info)
    assert result is True

    encoded_iri = urllib.parse.quote(ontology_iri, safe="")
    response = requests.get(
        f"{creds.server}/v2/ontologies/allentities/{encoded_iri}",
        timeout=10,
    )
    assert response.status_code == 200
    body = response.json()
    class_iri = f"{ontology_iri}#{CLASS_LOCAL_NAME}"
    classes = body.get("@graph", [])
    target = next((c for c in classes if c.get("@id") == class_iri), None)
    assert target is not None
    sub_class_of = target.get("rdfs:subClassOf", [])
    external_iris = [
        entry["@id"] for entry in (sub_class_of if isinstance(sub_class_of, list) else [sub_class_of]) if "@id" in entry
    ]
    assert "http://schema.org/Thing" in external_iris
