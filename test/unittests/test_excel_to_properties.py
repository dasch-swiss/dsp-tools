"""unit tests for excel to properties"""
import os
import unittest
import pandas as pd
import re
import json
import jsonpath_ng

from openpyxl import Workbook

from knora.dsplib.utils import excel_to_json_properties as e2j


class TestExcelToProperties(unittest.TestCase):

    def setUp(self) -> None:
        """Is executed before each test"""
        os.makedirs('testdata/tmp', exist_ok=True)

    def test_excel2json(self) -> None:
        excelfile = "testdata/Properties.xlsx"
        outfile = "testdata/tmp/_out_properties.json"
        e2j.properties_excel2json(excelfile, outfile)

        # read excel: skip all rows that lack one of the required values
        excel_df = pd.read_excel(excelfile)
        excel_df = excel_df[pd.notna(excel_df['name'])]
        excel_df = excel_df[[bool(re.search(r'\w', x)) for x in excel_df['name']]]
        excel_df = excel_df[pd.notna(excel_df['super'])]
        excel_df = excel_df[[bool(re.search(r'\w', x)) for x in excel_df['super']]]
        excel_df = excel_df[pd.notna(excel_df['object'])]
        excel_df = excel_df[[bool(re.search(r'\w', x)) for x in excel_df['object']]]
        excel_df = excel_df[pd.notna(excel_df['gui_element'])]
        excel_df = excel_df[[bool(re.search(r'\w', x)) for x in excel_df['gui_element']]]

        # extract infos from excel file
        excel_names = [s.strip() for s in excel_df['name']]
        excel_supers = [[x.strip() for x in s.split(',')] for s in excel_df['super']]
        excel_objects = [s.strip() for s in excel_df['object']]
        excel_labels_en = [s.strip() if isinstance(s, str) and re.search(r'\w', s) else '' for s in list(excel_df['en'])]
        excel_labels_de = [s.strip() if isinstance(s, str) and re.search(r'\w', s) else '' for s in list(excel_df['de'])]
        excel_labels_fr = [s.strip() if isinstance(s, str) and re.search(r'\w', s) else '' for s in list(excel_df['fr'])]
        excel_labels_it = [s.strip() if isinstance(s, str) and re.search(r'\w', s) else '' for s in list(excel_df['it'])]
        excel_comments_en = [s.strip() if isinstance(s, str) and re.search(r'\w', s) else '' for s in list(excel_df['comment_en'])]
        excel_comments_de = [s.strip() if isinstance(s, str) and re.search(r'\w', s) else '' for s in list(excel_df['comment_de'])]
        excel_comments_fr = [s.strip() if isinstance(s, str) and re.search(r'\w', s) else '' for s in list(excel_df['comment_fr'])]
        excel_comments_it = [s.strip() if isinstance(s, str) and re.search(r'\w', s) else '' for s in list(excel_df['comment_it'])]
        excel_gui_elements = [s.strip() for s in list(excel_df['gui_element'])]

        # read json file
        with open(outfile) as f:
            json_string = f.read()
            json_string = '{' + json_string + '}'
            json_file = json.loads(json_string)

        # extract infos from json file
        json_names = [match.value for match in jsonpath_ng.parse('$.properties[*].name').find(json_file)]
        json_supers = [match.value for match in jsonpath_ng.parse('$.properties[*].super').find(json_file)]
        json_objects = [match.value for match in jsonpath_ng.parse('$.properties[*].object').find(json_file)]
        json_labels = [match.value for match in jsonpath_ng.parse('$.properties[*].labels').find(json_file)]
        json_labels_en = [label.get('en', '').strip() for label in json_labels]
        json_labels_de = [label.get('de', '').strip() for label in json_labels]
        json_labels_fr = [label.get('fr', '').strip() for label in json_labels]
        json_labels_it = [label.get('it', '').strip() for label in json_labels]
        json_comments = [match.value for match in jsonpath_ng.parse('$.properties[*].comments').find(json_file)]
        json_comments_en = [comment.get('en', '').strip() for comment in json_comments]
        json_comments_de = [comment.get('de', '').strip() for comment in json_comments]
        json_comments_fr = [comment.get('fr', '').strip() for comment in json_comments]
        json_comments_it = [comment.get('it', '').strip() for comment in json_comments]
        json_gui_elements = [match.value for match in jsonpath_ng.parse('$.properties[*].gui_element').find(json_file)]

        # make checks
        self.assertListEqual(excel_names, json_names)
        self.assertListEqual(excel_supers, json_supers)
        self.assertListEqual(excel_objects, json_objects)
        self.assertListEqual(excel_labels_en, json_labels_en)
        self.assertListEqual(excel_labels_de, json_labels_de)
        self.assertListEqual(excel_labels_fr, json_labels_fr)
        self.assertListEqual(excel_labels_it, json_labels_it)
        self.assertListEqual(excel_comments_en, json_comments_en)
        self.assertListEqual(excel_comments_de, json_comments_de)
        self.assertListEqual(excel_comments_fr, json_comments_fr)
        self.assertListEqual(excel_comments_it, json_comments_it)
        self.assertListEqual(excel_gui_elements, json_gui_elements)

    def test_row_to_prop(self) -> None:
        wb = Workbook()
        ws = wb.create_sheet("Tabelle1")
        row = (
            "hasAnthroponym  ",
            " hasValue, dcterms:creator   ",
            " TextValue ",
            "anthroponym",
            "",
            "Anthroponyme",
            "",
            " A strange chance put me in possession of this journal.  ",
            "",
            "",
            "",
            " Richtext ",
            ""
        )
        for i, c in enumerate(row):
            ws.cell(row=2, column=i+1, value=c)
        properties_dict = e2j.row_to_prop(row)
        expected_dict = {
            "name": "hasAnthroponym",
            "super": [
                "hasValue",
                "dcterms:creator"
            ],
            "object": "TextValue",
            "labels": {
                "en": "anthroponym",
                "fr": "Anthroponyme",
            },
            "comments": {
                "en": "A strange chance put me in possession of this journal."
            },
            "gui_element": "Richtext"
        }
        self.assertDictEqual(properties_dict, expected_dict)


if __name__ == '__main__':
    unittest.main()
