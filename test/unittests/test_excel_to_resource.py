"""unit tests for excel to resource"""
import os
import unittest
import pandas as pd
import json
import jsonpath_ng
import re

from openpyxl import Workbook

from knora.dsplib.utils import excel_to_json_resources as e2j


class TestExcelToResource(unittest.TestCase):

    def setUp(self) -> None:
        """Is executed before each test"""
        os.makedirs('testdata/tmp', exist_ok=True)

    def test_excel2json(self) -> None:
        excelfile = "testdata/Resources.xlsx"
        outfile = "testdata/tmp/_out_res.json"
        e2j.resources_excel2json(excelfile, outfile)

        # read excel: skip all rows that lack one of the required values
        excel_df = pd.read_excel(excelfile)
        excel_df = excel_df[pd.notna(excel_df['name'])]
        excel_df = excel_df[[bool(re.search(r'[\wäöü]', x)) for x in excel_df['name']]]
        excel_df = excel_df[pd.notna(excel_df['super'])]
        excel_df = excel_df[[bool(re.search(r'[\wäöü]', x)) for x in excel_df['super']]]
        excel_first_class_df = pd.read_excel(excelfile, sheet_name=1)
        excel_first_class_df = excel_first_class_df[pd.notna(excel_first_class_df['Property'])]
        excel_first_class_df = excel_first_class_df[[bool(re.search(r'[\wäöü]', x)) for x in excel_first_class_df['Property'].astype(str)]]
        excel_first_class_df = excel_first_class_df[pd.notna(excel_first_class_df['Cardinality'])]
        excel_first_class_df = excel_first_class_df[[bool(re.search(r'[\wäöü]', x)) for x in excel_first_class_df['Cardinality'].astype(str)]]

        # extract infos from excel file
        excel_names = [s.strip() for s in excel_df['name']]
        excel_supers = [[x.strip() for x in s.split(',')] for s in excel_df['super']]
        excel_labels_en = [s.strip() if isinstance(s, str) and re.search(r'[\wäöü]', s) else '' for s in list(excel_df['en'])]
        excel_labels_de = [s.strip() if isinstance(s, str) and re.search(r'[\wäöü]', s) else '' for s in list(excel_df['de'])]
        excel_labels_fr = [s.strip() if isinstance(s, str) and re.search(r'[\wäöü]', s) else '' for s in list(excel_df['fr'])]
        excel_labels_it = [s.strip() if isinstance(s, str) and re.search(r'[\wäöü]', s) else '' for s in list(excel_df['it'])]
        excel_comments_en = [s.strip() if isinstance(s, str) and re.search(r'[\wäöü]', s) else '' for s in list(excel_df['comment_en'])]
        excel_comments_de = [s.strip() if isinstance(s, str) and re.search(r'[\wäöü]', s) else '' for s in list(excel_df['comment_de'])]
        excel_comments_fr = [s.strip() if isinstance(s, str) and re.search(r'[\wäöü]', s) else '' for s in list(excel_df['comment_fr'])]
        excel_comments_it = [s.strip() if isinstance(s, str) and re.search(r'[\wäöü]', s) else '' for s in list(excel_df['comment_it'])]
        excel_first_class_properties = [f':{s.strip()}' for s in excel_first_class_df['Property']]
        excel_first_class_cardinalities = [str(s).strip().lower() for s in excel_first_class_df['Cardinality']]

        # read json file
        with open(outfile) as f:
            json_string = f.read()
            json_string = '{' + json_string + '}'
            json_file = json.loads(json_string)

        # extract infos from json file
        json_names = [match.value for match in jsonpath_ng.parse('$.resources[*].name').find(json_file)]
        json_supers = [match.value for match in jsonpath_ng.parse('$.resources[*].super').find(json_file)]
        json_labels = [match.value for match in jsonpath_ng.parse('$.resources[*].labels').find(json_file)]
        json_labels_en = [label.get('en', '').strip() for label in json_labels]
        json_labels_de = [label.get('de', '').strip() for label in json_labels]
        json_labels_fr = [label.get('fr', '').strip() for label in json_labels]
        json_labels_it = [label.get('it', '').strip() for label in json_labels]
        json_comments = [match.value for match in jsonpath_ng.parse('$.resources[*].comments').find(json_file)]
        json_comments_en = [comment.get('en', '').strip() for comment in json_comments]
        json_comments_de = [comment.get('de', '').strip() for comment in json_comments]
        json_comments_fr = [comment.get('fr', '').strip() for comment in json_comments]
        json_comments_it = [comment.get('it', '').strip() for comment in json_comments]
        json_first_class_properties = [match.value for match in jsonpath_ng.parse('$.resources[0].cardinalities[*].propname').find(json_file)]
        json_first_class_cardinalities = [match.value for match in jsonpath_ng.parse('$.resources[0].cardinalities[*].cardinality').find(json_file)]

        # make checks
        self.assertListEqual(excel_names, json_names)
        self.assertListEqual(excel_supers, json_supers)
        self.assertListEqual(excel_labels_en, json_labels_en)
        self.assertListEqual(excel_labels_de, json_labels_de)
        self.assertListEqual(excel_labels_fr, json_labels_fr)
        self.assertListEqual(excel_labels_it, json_labels_it)
        self.assertListEqual(excel_comments_en, json_comments_en)
        self.assertListEqual(excel_comments_de, json_comments_de)
        self.assertListEqual(excel_comments_fr, json_comments_fr)
        self.assertListEqual(excel_comments_it, json_comments_it)
        self.assertListEqual(excel_first_class_properties, json_first_class_properties)
        self.assertListEqual(excel_first_class_cardinalities, json_first_class_cardinalities)

    def test_extract_row(self) -> None:
        wb = Workbook()
        ws_classes = wb.create_sheet("classes")
        res_name = "ClassA"
        row = (
            res_name,
            "Class A",
            "",
            "",
            "",
            "A comment on Class A",
            "",
            "",
            "",
            "Resource",
        )
        for i, c in enumerate(row):
            ws_classes.cell(row=2, column=i+1, value=c)
        ws_class_a = wb.create_sheet(res_name)
        ws_class_a["A2"] = "property1"
        ws_class_a["B2"] = "1"
        resource_dict = e2j._extract_row(row, wb)
        expected_dict = {
            'name': 'ClassA',
            'labels': {
                'en': 'Class A'
            },
            'comments': {
                'en': 'A comment on Class A'
            },
            'super': ['Resource'],
            'cardinalities': [{
                'propname': ':property1',
                'cardinality': '1',
                'gui_order': 1
            }]}
        self.assertDictEqual(resource_dict, expected_dict)


if __name__ == '__main__':
    unittest.main()
